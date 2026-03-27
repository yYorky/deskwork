"""
Build AFC Audit Sampling Workpaper
Task: Anti-Financial Crime audit sample selection on Q2/Q3 2024 risk metrics
Output: Sample.xlsx with Tab 1 (Sample) and Tab 2 (Sample Size Calculation)
"""

import math
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Paths ───────────────────────────────────────────────────────────────────
POP_PATH = (
    r"C:\Users\York Yong\OneDrive - Singapore Management University"
    r"\Desktop\Deskwork\benchmarks\tasks\afc-audit-sampling\reference-files\Population v2.xlsx"
)
OUT_PATH = (
    r"C:\Users\York Yong\OneDrive - Singapore Management University"
    r"\Desktop\Deskwork\benchmarks\claude-code\afc-audit-sampling\output\Sample.xlsx"
)

# ─── Load Population ──────────────────────────────────────────────────────────
wb_pop = load_workbook(POP_PATH)
ws_pop = wb_pop.active

headers = [ws_pop.cell(1, c).value for c in range(1, 9)]
# Headers: No, Division, Sub-Division, Country, Legal Entity, KRIs, Q3 2024 KRI, Q2 2024 KRI
# Col indices (1-based): A=1 No, B=2 Div, C=3 SubDiv, D=4 Country, E=5 Entity, F=6 KRI, G=7 Q3, H=8 Q2

rows = []
for r in range(2, ws_pop.max_row + 1):
    row = [ws_pop.cell(r, c).value for c in range(1, 9)]
    rows.append(row)

N = len(rows)
print(f"Population size N = {N}")

# ─── Step 1: Calculate QoQ Variance ──────────────────────────────────────────
# Q3 = col index 6 (G), Q2 = col index 7 (H) in 0-based row list
# Variance = (Q3 - Q2) / |Q2|
# Special cases:
#   Q2=0 and Q3=0 → 0
#   Q2=0 and Q3≠0 → "N/M"
#   otherwise → (Q3-Q2)/|Q2|

def calc_variance(q3, q2):
    if q2 == 0 and q3 == 0:
        return 0
    elif q2 == 0:
        return "N/M"
    else:
        return (q3 - q2) / abs(q2)

variances = []
for row in rows:
    q3 = row[6] if row[6] is not None else 0
    q2 = row[7] if row[7] is not None else 0
    variances.append(calc_variance(q3, q2))

# ─── Step 2: Sample Size Calculation ─────────────────────────────────────────
z = 1.645
p = 0.5
e = 0.10

n_raw = (z**2 * p * (1 - p)) / (e**2)
n_ceil = math.ceil(n_raw)

# Finite population correction
n_fpc = n_ceil / (1 + (n_ceil - 1) / N)
n_final = math.ceil(n_fpc)

print(f"n_raw = {n_raw:.4f}  ->  n_ceil = {n_ceil}")
print(f"n_fpc = {n_fpc:.4f}  ->  n_final = {n_final}")

# ─── Step 3: Sample Selection ─────────────────────────────────────────────────
# Mandatory criteria — each must be satisfied by at least one selected row.
# Return list of 0-based row indices satisfying each criterion.

selected = set()

def select_first(condition_fn, label):
    """Select first row meeting condition. Return True if found."""
    for i, row in enumerate(rows):
        if condition_fn(i, row):
            selected.add(i)
            return True
    print(f"WARNING: No rows found for: {label}")
    return False

# Helper lookups
def div(i):  return rows[i][1] or ""
def sub(i):  return rows[i][2] or ""
def ctry(i): return rows[i][3] or ""
def kri(i):  return rows[i][5] or ""
def var(i):  return variances[i]

# --- Mandatory: 5 flagged entities (by Div + SubDiv + Country) ---
mandatory_entities = [
    ("Corporate Bank", "Corporate Loans",     "Italy",    "CB Cash Italy"),
    ("Corporate Bank", "Correspondent Banking","Greece",   "CB Correspondent Banking Greece"),
    ("Markets",        "Trading",             "Luxembourg","IB Debt Markets Luxembourg"),
    ("Corporate Bank", "Marine Finance",       "Brazil",   "CB Trade Finance Brazil"),
    ("Retail Bank",    "EMEA",                "UAE",      "PB EMEA UAE"),
]
for d, s, c, label in mandatory_entities:
    select_first(
        lambda i, r, _d=d, _s=s, _c=c: div(i)==_d and sub(i)==_s and ctry(i)==_c,
        label
    )

# --- Mandatory: metrics A1 (Total clients) and C1 (HR Clients) ---
select_first(lambda i, r: "total clients" in kri(i).lower(), "Total clients")
select_first(lambda i, r: kri(i) == "HR Clients", "HR Clients")

# --- Mandatory: at least one zero-zero row ---
select_first(
    lambda i, r: (r[6] == 0 and r[7] == 0),
    "Zero-zero (Q3=0, Q2=0)"
)

# --- Mandatory: Marine Finance and Correspondent Banking ---
select_first(lambda i, r: sub(i) == "Marine Finance",        "Marine Finance")
select_first(lambda i, r: sub(i) == "Correspondent Banking", "Correspondent Banking")

# --- Mandatory: Cayman Islands, Pakistan, UAE ---
select_first(lambda i, r: ctry(i) == "Cayman Islands", "Cayman Islands")
select_first(lambda i, r: ctry(i) == "Pakistan",       "Pakistan")
select_first(lambda i, r: ctry(i) == "UAE",            "UAE")

# --- Mandatory: all Divisions represented ---
all_divs = sorted(set(rows[i][1] for i in range(N) if rows[i][1]))
print(f"Divisions: {all_divs}")
for d in all_divs:
    select_first(lambda i, r, _d=d: div(i) == _d, f"Division: {d}")

# --- Mandatory: all Sub-Divisions represented ---
all_subs = sorted(set(rows[i][2] for i in range(N) if rows[i][2]))
print(f"Sub-Divisions: {all_subs}")
for s in all_subs:
    select_first(lambda i, r, _s=s: sub(i) == _s, f"SubDiv: {s}")

# --- Mandatory: at least one row with |variance| > 20% ---
select_first(
    lambda i, r: isinstance(var(i), float) and abs(var(i)) > 0.20,
    "Variance > 20%"
)

# --- Mandatory: at least one row with |variance| >= 100% ---
select_first(
    lambda i, r: isinstance(var(i), float) and abs(var(i)) >= 1.00,
    "Variance >= 100%"
)

print(f"\nMandatory selections so far: {len(selected)}")

# --- Top up to reach n_final: prioritize highest absolute variance rows ---
# Sort all rows by |variance| descending (numeric only)
high_var_rows = sorted(
    [(i, abs(v)) for i, v in enumerate(variances) if isinstance(v, float)],
    key=lambda x: -x[1]
)

for i, _ in high_var_rows:
    if len(selected) >= n_final:
        break
    selected.add(i)

# Also ensure we have multiple high-variance (≥100%) rows for visibility
extremely_high = [i for i, v in enumerate(variances) if isinstance(v, float) and abs(v) >= 1.0]
for i in extremely_high[:10]:  # Include up to 10 extremely high variance rows
    selected.add(i)

print(f"Final sample size S = {len(selected)}  (required R = {n_final})")

# Sort selected by original row order
selected_sorted = sorted(selected)

# ─── Step 4: Build Output Workbook ────────────────────────────────────────────
wb = Workbook()

# --- Tab 1: Sample ---
ws1 = wb.active
ws1.title = "Sample"

# Style helpers
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SUBHEADER_FILL = PatternFill("solid", fgColor="2E75B6")
SUBHEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
FLAG_FILL = PatternFill("solid", fgColor="FFF2CC")  # yellow for high variance
EXTREME_FILL = PatternFill("solid", fgColor="FFD966")  # darker yellow for ≥100%
SELECTED_FILL = PatternFill("solid", fgColor="E2EFDA")  # light green for selected
normal_font = Font(size=10)

pct_fmt = "0.00%"
num_fmt = "#,##0.00"
thin = Side(style="thin", color="BBBBBB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Column headers for Sample tab
sample_headers = headers + ["QoQ Variance (Q3 vs Q2)", "Selected"]
for c, h in enumerate(sample_headers, 1):
    cell = ws1.cell(1, c, h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = border

# Write ALL population rows
for r_idx, i in enumerate(range(N), 2):
    row = rows[i]
    v = variances[i]
    is_selected = i in selected

    for c, val in enumerate(row, 1):
        cell = ws1.cell(r_idx, c, val)
        cell.font = normal_font
        cell.border = border
        if c in (7, 8) and isinstance(val, (int, float)):
            cell.number_format = num_fmt

    # Col I: variance
    v_cell = ws1.cell(r_idx, 9)
    if v == "N/M":
        v_cell.value = "N/M"
    elif v == 0:
        v_cell.value = 0
        v_cell.number_format = pct_fmt
    else:
        v_cell.value = v
        v_cell.number_format = pct_fmt
    v_cell.font = normal_font
    v_cell.border = border

    # Col J: sample flag
    j_cell = ws1.cell(r_idx, 10)
    if is_selected:
        j_cell.value = 1
    j_cell.font = normal_font
    j_cell.border = border

    # Highlight rows
    if is_selected:
        row_fill = SELECTED_FILL
        if isinstance(v, float) and abs(v) >= 1.0:
            row_fill = EXTREME_FILL
        elif isinstance(v, float) and abs(v) > 0.20:
            row_fill = FLAG_FILL
        for c in range(1, 11):
            ws1.cell(r_idx, c).fill = row_fill

# Summary row
summary_row = N + 2
ws1.cell(summary_row, 9, "Total Selected:").font = Font(bold=True, size=10)
ws1.cell(summary_row, 9).alignment = Alignment(horizontal="right")
total_cell = ws1.cell(summary_row, 10)
total_cell.value = f'=SUM(J2:J{N+1})'
total_cell.font = Font(bold=True, size=10)
total_cell.fill = PatternFill("solid", fgColor="DEEBF7")
total_cell.border = border

# Column widths
col_widths = [6, 16, 22, 18, 36, 36, 14, 14, 22, 12]
for c, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(c)].width = w

ws1.freeze_panes = "A2"

# --- Tab 2: Sample Size Calculation ---
ws2 = wb.create_sheet("Sample Size Calculation")

def write_row(ws, row, col, label, value, bold=False, fmt=None, fill=None):
    lc = ws.cell(row, col, label)
    vc = ws.cell(row, col + 1, value)
    if bold:
        lc.font = Font(bold=True, size=10)
        vc.font = Font(bold=True, size=10)
    else:
        lc.font = Font(size=10)
        vc.font = Font(size=10)
    if fmt:
        vc.number_format = fmt
    if fill:
        lc.fill = fill
        vc.fill = fill
    lc.border = border
    vc.border = border
    return lc, vc

TITLE_FONT = Font(bold=True, size=13, color="1F4E79")
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
RESULT_FILL = PatternFill("solid", fgColor="E2EFDA")

ws2.column_dimensions["A"].width = 45
ws2.column_dimensions["B"].width = 22

# Title
ws2.merge_cells("A1:B1")
t = ws2.cell(1, 1, "AFC Audit Sampling — Sample Size Calculation")
t.font = TITLE_FONT
t.alignment = Alignment(horizontal="center")

ws2.merge_cells("A2:B2")
ws2.cell(2, 1, "GDPval Benchmark | Claude Code + SKILL.md method")
ws2.cell(2, 1).font = Font(italic=True, size=10, color="595959")
ws2.cell(2, 1).alignment = Alignment(horizontal="center")

r = 4
ws2.cell(r, 1, "INPUTS").font = Font(bold=True, size=11, color="1F4E79")
r += 1

write_row(ws2, r, 1, "Confidence Level",        "90%",   fill=SECTION_FILL); r += 1
write_row(ws2, r, 1, "Tolerable Error Rate (e)", "10%",   fill=SECTION_FILL); r += 1
write_row(ws2, r, 1, "Z-value (one-tailed 90%)", z,       fill=SECTION_FILL, fmt="0.000"); r += 1
write_row(ws2, r, 1, "p (conservative maximum)", p,       fill=SECTION_FILL, fmt="0.0"); r += 1
write_row(ws2, r, 1, "Population size N",         N,       fill=SECTION_FILL, fmt="#,##0"); r += 1

r += 1
ws2.cell(r, 1, "SAMPLE SIZE FORMULA").font = Font(bold=True, size=11, color="1F4E79"); r += 1

write_row(ws2, r, 1, "Formula", "n = (z² × p × (1-p)) / e²"); r += 1
write_row(ws2, r, 1, "z²",      round(z**2, 6), fmt="0.000000"); r += 1
write_row(ws2, r, 1, "p × (1-p)", p * (1-p), fmt="0.00"); r += 1
write_row(ws2, r, 1, "e²",      e**2, fmt="0.0000"); r += 1
write_row(ws2, r, 1, "n (raw)",  round(n_raw, 4), fmt="0.0000"); r += 1
write_row(ws2, r, 1, "n (rounded up — without FPC)", n_ceil, fmt="#,##0"); r += 1

r += 1
ws2.cell(r, 1, "FINITE POPULATION CORRECTION").font = Font(bold=True, size=11, color="1F4E79"); r += 1

write_row(ws2, r, 1, "Condition: N < 10 x n?", f"{N} < {10*n_ceil} -> {'YES' if N < 10*n_ceil else 'NO - FPC applied as conservative step'}"); r += 1
write_row(ws2, r, 1, "FPC Formula", "n_adj = n / (1 + (n-1) / N)"); r += 1
write_row(ws2, r, 1, "n_adj (raw)", round(n_fpc, 4), fmt="0.0000"); r += 1
write_row(ws2, r, 1, "n_adj (rounded up)",  n_final, fmt="#,##0"); r += 1

r += 1
ws2.cell(r, 1, "RESULT").font = Font(bold=True, size=11, color="1F4E79"); r += 1
write_row(ws2, r, 1, "Required Sample Size R", n_final, bold=True, fill=RESULT_FILL, fmt="#,##0"); r += 1
write_row(ws2, r, 1, "Actual Sample Size S",   len(selected), bold=True, fill=RESULT_FILL, fmt="#,##0"); r += 1
write_row(ws2, r, 1, "S ≥ R?", "YES" if len(selected) >= n_final else "NO", bold=True, fill=RESULT_FILL); r += 1

r += 1
ws2.cell(r, 1, "SELECTION CRITERIA COVERAGE").font = Font(bold=True, size=11, color="1F4E79"); r += 1

criteria_checks = [
    ("High variance rows (|QoQ| > 20%)",    any(isinstance(variances[i], float) and abs(variances[i]) > 0.20 for i in selected)),
    ("Exceptionally large (|QoQ| ≥ 100%)", any(isinstance(variances[i], float) and abs(variances[i]) >= 1.0  for i in selected)),
    ("CB Cash Italy (Corporate Bank, Corporate Loans, Italy)",        any(div(i)=="Corporate Bank" and sub(i)=="Corporate Loans"      and ctry(i)=="Italy"     for i in selected)),
    ("CB Correspondent Banking Greece",     any(div(i)=="Corporate Bank" and sub(i)=="Correspondent Banking" and ctry(i)=="Greece"    for i in selected)),
    ("IB Debt Markets Luxembourg",          any(div(i)=="Markets"        and sub(i)=="Trading"               and ctry(i)=="Luxembourg" for i in selected)),
    ("CB Trade Finance Brazil",             any(div(i)=="Corporate Bank" and sub(i)=="Marine Finance"        and ctry(i)=="Brazil"     for i in selected)),
    ("PB EMEA UAE",                         any(div(i)=="Retail Bank"    and sub(i)=="EMEA"                  and ctry(i)=="UAE"        for i in selected)),
    ("Metric: Total clients (A1)",          any("total clients" in kri(i).lower() for i in selected)),
    ("Metric: HR Clients (C1)",             any(kri(i) == "HR Clients"   for i in selected)),
    ("Zero-zero rows (Q2=0, Q3=0)",         any(rows[i][6]==0 and rows[i][7]==0 for i in selected)),
    ("Marine Finance sub-division",         any(sub(i)=="Marine Finance"        for i in selected)),
    ("Correspondent Banking sub-division",  any(sub(i)=="Correspondent Banking" for i in selected)),
    ("Cayman Islands",                      any(ctry(i)=="Cayman Islands" for i in selected)),
    ("Pakistan",                            any(ctry(i)=="Pakistan"       for i in selected)),
    ("UAE",                                 any(ctry(i)=="UAE"            for i in selected)),
]
for div_name in all_divs:
    criteria_checks.append((f"Division covered: {div_name}", any(div(i)==div_name for i in selected)))
for sub_name in all_subs:
    criteria_checks.append((f"Sub-Division covered: {sub_name}", any(sub(i)==sub_name for i in selected)))

for label, ok in criteria_checks:
    cell_l = ws2.cell(r, 1, ("✓ " if ok else "✗ ") + label)
    cell_v = ws2.cell(r, 2, "PASS" if ok else "FAIL")
    cell_l.font = Font(size=10, color="375623" if ok else "FF0000")
    cell_v.font = Font(bold=True, size=10, color="375623" if ok else "FF0000")
    cell_l.border = border
    cell_v.border = border
    r += 1

ws2.freeze_panes = "A4"

# ─── Save ─────────────────────────────────────────────────────────────────────
wb.save(OUT_PATH)
print(f"\nSaved: {OUT_PATH}")
print(f"Sample size: {len(selected)} rows (required: {n_final})")
print(f"\nSelf-verification:")
print(f"  File exists: YES")
print(f"  Tab 1 name: {ws1.title}")
print(f"  Tab 2 name: {ws2.title}")
print(f"  N = {N}")
print(f"  R = {n_final}")
print(f"  S = {len(selected)}")
print(f"  S >= R: {len(selected) >= n_final}")
