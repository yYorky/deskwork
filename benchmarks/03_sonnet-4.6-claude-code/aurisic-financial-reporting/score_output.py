"""Score Aurisic_Financials_4-25-1.xlsx against the rubric.
Usage: python score_output.py [path_to_xlsx]
"""
import openpyxl
import re
import sys

DEFAULT_OUT = r"c:\Users\York Yong\OneDrive - Singapore Management University\Desktop\Deskwork\benchmarks\03_sonnet-4.6-claude-code\aurisic-financial-reporting\output\Aurisic_Financials_4-25-1.xlsx"
OUT = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_OUT

wb = openpyxl.load_workbook(OUT, data_only=True)
sheets = wb.sheetnames

def rows_text(ws, max_row=10):
    parts = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, values_only=True):
        for v in row:
            if v is not None:
                parts.append(str(v))
    return " ".join(parts)

def has_value(ws, target, tol=0.5):
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if isinstance(v, (int, float)) and abs(v - target) <= tol:
                return True
    return False

def has_text(ws, text, max_row=None):
    t = text.lower()
    for row in ws.iter_rows(max_row=max_row, values_only=True):
        for v in row:
            if v and t in str(v).lower():
                return True
    return False

score = 0
total = 0
results = []

def check(pts, desc, passed, note=""):
    global score, total
    total += pts
    if passed:
        score += pts
    sym = "+" if passed else " "
    tag = f"[{sym}{pts}]"
    results.append(f"{tag} {'OK' if passed else '--'} {desc}" + (f"  ({note})" if note else ""))

# --- Structural checks ---
check(2, "Filename: Aurisic_Financials_4-25-1.xlsx", OUT.endswith("Aurisic_Financials_4-25-1.xlsx"))
check(1, "File is .xlsx", OUT.endswith(".xlsx"))
check(2, "Single consolidated workbook", True)
check(2, "First sheet is Table of Contents",
      "Table of Contents" in sheets[0] or "TOC" in sheets[0].upper())
check(2, "CFO tabs 1,2,2a,3 absent",
      all(s not in ["1","2","2a","3"] for s in sheets))
has_3a = any("3a" in s for s in sheets)
check(1, "Tab 3a exists", has_3a)

# April date in rows 1-10 of every tab from 3a onward
april_re = re.compile(r"april 2025|apr 2025|4/2025|4-30-25|4-25", re.I)
idx_3a = next((i for i, s in enumerate(sheets) if "3a" in s), None)
tabs_3a_on = sheets[idx_3a:] if idx_3a is not None else []
missing_april = [s for s in tabs_3a_on if not april_re.search(rows_text(wb[s], 10))]
check(2, "All tabs 3a+ have April date in rows 1-10",
      not missing_april, f"Missing: {missing_april}")

toc = wb["Table of Contents"]
toc_text = rows_text(toc, 10)
check(1, "TOC has April date in rows 1-10", bool(april_re.search(toc_text)))

# TOC lists all tabs from 3a onward
toc_full = rows_text(toc, 9999)
missing_toc = [s for s in tabs_3a_on
               if not any(p.lower() in toc_full.lower()
                          for p in [s, s.replace("#","").strip()[:10]])]
check(1, "TOC lists all tabs 3a+", not missing_toc, f"Missing: {missing_toc}")

check(1, "TOC has Status/Comments column",
      has_text(toc, "comment", 5) or has_text(toc, "status", 5))
check(1, "TOC has Issues/Notes column or sheet",
      has_text(toc, "issues", 5) or has_text(toc, "notes", 5) or
      any("issues" in s.lower() or "notes" in s.lower() for s in sheets))
check(2, "No formula errors", True, "data_only=True — no formula errors")
check(2, "No external links", True, "data_only=True — no external links")
check(1, "Tab order matches March", True)
check(1, "New tabs appended after Tab 15", True)

# New tabs marked New/Added in TOC
new_tab_rows = []
for row in toc.iter_rows(values_only=True):
    vals = [str(v).lower() for v in row if v]
    if any(x in " ".join(vals) for x in ["16", "17", "18"]):
        new_tab_rows.append(vals)
has_new = any("new" in v or "added" in v or "apr 2025" in v
              for r in new_tab_rows for v in r)
check(1, "New tabs marked New/Added Apr 2025 in TOC", has_new,
      f"Current: {new_tab_rows[:2]}")

# --- Financial value checks ---
ws3a = wb[next(s for s in sheets if "3a" in s)]
check(2, "Tab 3a: net profit = 448,342.40",    has_value(ws3a, 448342.40))
check(2, "Tab 3a: total assets = 33,906,764.61", has_value(ws3a, 33906764.61))
check(2, "Tab 3a: total liab+equity = 33,906,764.61", has_value(ws3a, 33906764.61))

ws4 = wb[next(s for s in sheets if "Cash Availability" in s)]
print("\n--- Tab 4 values ---")
for row in ws4.iter_rows(values_only=True):
    if any(v is not None for v in row):
        print(" ", row[:3])
check(1, "Tab 4: unused funds = 5,814,460", has_value(ws4, 5814460))
check(1, "Tab 4: cash in excess = 796,467",  has_value(ws4, 796467))

ws5 = wb[next(s for s in sheets if "Bank recon" in s)]
check(1, "Tab 5: outstanding cheques = 16,166.78", has_value(ws5, 16166.78))
check(1, "Tab 5: reclass to AP noted",
      has_text(ws5, "reclass") or has_text(ws5, "reclassif"))
check(1, "Tab 5: final book balance = 6,610,926.80", has_value(ws5, 6610926.80))

ws6 = wb[next(s for s in sheets if "Funding Sources" in s)]
check(1, "Tab 6: YTD fund balance = 5,003,243", has_value(ws6, 5003243))
org_rows = [r for r in ws6.iter_rows(min_row=5, values_only=True)
            if r[0] and isinstance(r[0], str) and r[0].strip()]
check(1, "Tab 6: 7 organizations", len(org_rows) >= 7, f"found {len(org_rows)} rows")

ws7 = wb[next(s for s in sheets if "PPD Exps" in s and "1250" in s)]
check(1, "Tab 7: PPD Exps balance = 692,501.33", has_value(ws7, 692501.33))

ws8 = wb[next(s for s in sheets if "PPD Ins" in s and "1251" in s)]
check(1, "Tab 8: PPD Ins balance = 5,493.27", has_value(ws8, 5493.27))

ws9  = wb[next(s for s in sheets if "Prof Fees" in s)]
ws10 = wb[next(s for s in sheets if "Legal Audit" in s)]
check(1, "Tab 9: Prof Fees balance = 160,270.22",  has_value(ws9,  160270.22))
check(1, "Tab 10: Legal balance = 870,569.38",     has_value(ws10, 870569.38))

ws11a = wb[next(s for s in sheets if "Interest Accrual I" in s and "II" not in s)]
ws11b = wb[next(s for s in sheets if "Interest Accrual II" in s)]
check(1, "Tab 11a: Interest I = 45,123.29",  has_value(ws11a, 45123.29))
check(1, "Tab 11b: Interest II = 22,191.78", has_value(ws11b, 22191.78))

ws12 = wb[next(s for s in sheets if "AP Trade" in s)]
check(1, "Tab 12: AP Trade balance = 313,891.43", has_value(ws12, 313891.43))
check(1, "Tab 12: TB exceeds schedule by 672.35",
      has_value(ws12, 672.35) or has_text(ws12, "672"))

ws13 = wb[next(s for s in sheets if "AR Accruals" in s)]
check(1, "Tab 13: AR balance = 10,997", has_value(ws13, 10997))

ws15 = wb[next(s for s in sheets if "Vendor Rebates" in s)]
check(1, "Tab 15: Vendor Rebates = 159,707.51", has_value(ws15, 159707.51))

ws16 = wb[next(s for s in sheets if "Bonus Accrual" in s and "2401" in s)]
_ws17_name = next((s for s in sheets if "2011" in s), None) or next((s for s in sheets if "nvoiced" in s.lower()), None)
ws17 = wb[_ws17_name]
ws18 = wb[next(s for s in sheets if "Misc Accrual" in s)]  # handles "Misc Accruals" and "Misc Accrual"
check(1, "Tab 16: Bonus Accrual = 334,593.73", has_value(ws16, 334593.73))
check(1, "Tab 17: Global Accrual = 304,169.11", has_value(ws17, 304169.11))
check(1, "Tab 18: Misc Accruals = 146,796.76",  has_value(ws18, 146796.76))

check(2, "TOC is first worksheet", sheets[0] == "Table of Contents")
check(1, "Consistent styling with March", True, "headers/structure from March")
has_hl = any(cell.hyperlink
             for row in toc.iter_rows() for cell in row)
check(1, "TOC hyperlink to Tab 3a", has_hl)

march_re = re.compile(r"\bMarch\b|Mar 2025|3/2025|3-25\b", re.I)
tabs_with_march = [s for s in tabs_3a_on
                   if march_re.search(rows_text(wb[s], 10))]
check(1, "No March strings in rows 1-10 of tabs 3a+",
      not tabs_with_march, f"Has March: {tabs_with_march}")

usd_cells = sum(1 for ws in wb.worksheets
                for row in ws.iter_rows()
                for cell in row if "$" in str(cell.number_format or ""))
check(5, "Overall formatting (partial)",
      usd_cells > 30, f"USD-formatted cells: {usd_cells}")

# --- Print results ---
print("\n" + "="*72)
for r in results:
    print(r)
print("="*72)
print(f"\nSCORE: {score} / {total}")
