#!/usr/bin/env python3
"""Build Aurisic_Financials_4-25-1.xlsx — April 2025 month-end package for Aurisic."""

import openpyxl
from openpyxl.styles import Font, PatternFill
import re
import os

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REF = r"c:\Users\York Yong\OneDrive - Singapore Management University\Desktop\Deskwork\benchmarks\tasks\aurisic-financial-reporting\reference-files"
OUT_DIR = os.path.join(SCRIPT_DIR, "output")
OUT_FILE = os.path.join(OUT_DIR, "Aurisic_Financials_4-25-1.xlsx")
os.makedirs(OUT_DIR, exist_ok=True)

FMT_USD = '$#,##0.00'


def flag_cell(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=f"FLAG: {text}")
    c.font = Font(bold=True, color="FF0000")
    c.fill = PatternFill("solid", fgColor="FFFF00")


def copy_values(src, dst):
    for row in src.iter_rows(values_only=True):
        dst.append(list(row))


def add_tab(fname, src_sheet, dst_name, wb_out):
    wb = openpyxl.load_workbook(os.path.join(REF, fname), data_only=True)
    ws_dst = wb_out.create_sheet(dst_name)
    copy_values(wb[src_sheet], ws_dst)
    wb.close()
    return ws_dst


# ── Parse April Trial Balance ──────────────────────────────────────────────────
def parse_tb(path):
    rows = []
    pat = re.compile(
        r"^(\d{4})-(\d{4})\s{2,}(.+?)\s{2,}([\d,]+\.?\d*[-]?)\s+([\d,]+\.?\d*[-]?)\s+([a-z])\s*$"
    )
    with open(path, encoding="utf-8", errors="replace") as f:
        for line in f:
            m = pat.match(line)
            if m:
                def n(s):
                    s = s.replace(",", "")
                    return -float(s[:-1]) if s.endswith("-") else float(s)
                rows.append((int(m[1]), int(m[2]), m[3].strip(), n(m[4]), n(m[5]), m[6]))
    return rows


tb = parse_tb(os.path.join(REF, "Aurisic_Final_TB_4-25-1.txt"))
print(f"TB rows parsed: {len(tb)}")


def ytd(acct, div=10):
    for r in tb:
        if r[0] == div and r[1] == acct:
            return r[4]
    return 0.0


# ── Open March template ────────────────────────────────────────────────────────
wb_mar = openpyxl.load_workbook(
    os.path.join(REF, "Aurisic_Financials_3-25-1.xlsx"), data_only=True
)

# ── Create output workbook ─────────────────────────────────────────────────────
wb_out = openpyxl.Workbook()
del wb_out[wb_out.sheetnames[0]]

# ══════════════════════════════════════════════════════════════════════════════
# TAB 0: Table of Contents
# ══════════════════════════════════════════════════════════════════════════════
ws0 = wb_out.create_sheet("Table of Contents")
for row in wb_mar["Table of Contents"].iter_rows(values_only=True):
    new = []
    for v in row:
        if isinstance(v, str):
            v = (v
                 .replace("March 2025", "April 2025")
                 .replace("3-31-25", "4-30-25")
                 .replace("2-28-25", "4-30-25")
                 .replace("March", "April")
                 .replace("3-25", "4-25"))
        new.append(v)
    ws0.append(new)
ws0.append(["16", "#16) Bonus Accrual #2401", "COMPLETE"])
ws0.append(["17", "#17) Global Accrual #2011", "COMPLETE"])
ws0.append(["18", "#18) Misc Accruals #2410", "COMPLETE"])
print("Tab 0 (TOC): done")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 3a: TB Convert 4-30-25
# ══════════════════════════════════════════════════════════════════════════════
ws3a = wb_out.create_sheet("#3a) TB convert 4-30-25")
hdr = ["Div #", "Acct #", "Account  Name", "MTD Amt", "YTD Amt", "code"]
ws3a.append(hdr)
for c in ws3a[1]:
    c.font = Font(bold=True)
for div, acct, name, mtd, yt, code in tb:
    r = ws3a.max_row + 1
    ws3a.append([div, acct, name, mtd, yt, code])
    ws3a.cell(r, 4).number_format = FMT_USD
    ws3a.cell(r, 5).number_format = FMT_USD
print(f"Tab 3a (TB): {len(tb)} rows loaded")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 4: Cash Availability Status
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb_out.create_sheet("#4) Cash Availability Status")
ic = {
    "Notes Rec - Aurisic UK": ytd(1651),
    "Notes Rec - Aurisic MX": ytd(1652),
    "Notes Rec - Aurisic CA": ytd(1653),
    "Notes Rec - Aurisic JP": ytd(1654),
    "Notes Rec - Aurisic DE": ytd(1656),
    "Notes Rec - Aurisic AU": ytd(1663),
    "Notes Rec - Aurisic BR": ytd(1664),
}
total_ic = sum(ic.values())
loan_bal = abs(ytd(2600)) + abs(ytd(2601))  # Loan I $18.3M + Loan II $13.5M
total_cash = ytd(1023) + ytd(1024)

ws4.append(["Aurisic Companies"])
ws4["A1"].font = Font(bold=True)
ws4.append(["Cash Availability from Sources"])
ws4.append(["as of 4-30-25"])
ws4.append(["Description", "Amount"])
ws4["A4"].font = Font(bold=True)
ws4["B4"].font = Font(bold=True)
ws4.append(["Good Insurance Co Loan Balance", loan_bal])
ws4.append(["  Loan I (acct 2600 - LOC)", abs(ytd(2600))])
ws4.append(["  Loan II (acct 2601 - Other LOC)", abs(ytd(2601))])
ws4.append(["Intercompany Loans:"])
for nm, v in ic.items():
    ws4.append([nm, v])
ws4.append(["Total Intercompany Loans", total_ic])
ws4.append(["Unused Funds from Good Insurance Co", loan_bal - total_ic])
ws4.append(["Cash Balance 4-30-25", total_cash])
for row in ws4.iter_rows(min_row=5, max_col=2):
    if isinstance(row[1].value, (int, float)):
        row[1].number_format = FMT_USD
print("Tab 4 (Cash Availability): done")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 5: Bank Recon 4-30-25
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb_out.create_sheet("#5) Bank recon 4-30-25")
wb_ck = openpyxl.load_workbook(
    os.path.join(REF, "Outstanding_CKs_4-30-25-1.xlsx"), data_only=True
)
copy_values(wb_ck["#6) Bank recon 4-30-25"], ws5)
wb_ck.close()
gl_cash = ytd(1024)
os_cks = 16166.78
bank_bal = gl_cash + os_cks
ws5.cell(4, 3, bank_bal).number_format = FMT_USD
flag_cell(
    ws5, 4, 5,
    f"Bank balance derived from GL cash (acct 1024) YTD ${gl_cash:,.2f} + "
    f"outstanding checks ${os_cks:,.2f} = ${bank_bal:,.2f}. "
    "Confirm against actual April bank statement before finalising."
)
print("Tab 5 (Bank Recon): done")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 6: Aurisic Funding Sources
# ══════════════════════════════════════════════════════════════════════════════
ws6 = wb_out.create_sheet("#6) Aurisic Funding Sources")
for row in wb_mar["#6) Aurisic Funding Sources"].iter_rows(values_only=True):
    new = []
    for v in row:
        if isinstance(v, str):
            v = (v
                 .replace("3-31-2025", "4-30-2025")
                 .replace("3-31-25", "4-30-25")
                 .replace("March", "April"))
        new.append(v)
    ws6.append(new)
flag_cell(
    ws6, 2, 4,
    "YTD Actual column not updated — April subsidiary revenue data not provided "
    "in source files. Notify CFO before finalising."
)
print("Tab 6 (Funding Sources): done (flagged)")

# ══════════════════════════════════════════════════════════════════════════════
# TABs 7 & 8: Prepaid Expenses and Insurance
# ══════════════════════════════════════════════════════════════════════════════
add_tab("PPD1250-1.xlsx", "PPD Exps #1250 (2025)", "#7) PPD Exps #1250 (2025)", wb_out)
add_tab("PPD1251-1.xlsx", "PPD Ins #1251 (2025)",  "#8) PPD Ins #1251 (2025)",  wb_out)
print("Tabs 7 & 8 (PPD): done")

# ══════════════════════════════════════════════════════════════════════════════
# TABs 9 & 10: GL Dump Reports
# ══════════════════════════════════════════════════════════════════════════════
add_tab("Prof_Fee_Dump-1.xlsx", "Prof Fees Dump", "#9) Prof Fees Accrual #2404",    wb_out)
add_tab("Legal_Dump-1.xlsx",   "Legal Dump",      "#10) Legal Audit Expense #6200", wb_out)
print("Tabs 9 & 10 (GL Dumps): done")

# ══════════════════════════════════════════════════════════════════════════════
# TABs 11a & 11b: Interest Accrual Schedules
# ══════════════════════════════════════════════════════════════════════════════
add_tab("Good Insurance Co - Loan.xlsx",    "OMAHA",   "#11a) Interest Accrual I",  wb_out)
add_tab("Good Insurance Co - Loan II.xlsx", "OMAHAII", "#11b) Interest Accrual II", wb_out)
print("Tabs 11a & 11b (Interest): done")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 12: AP Trade #2000
# AP_TB-1.xlsx not present — derive from TB
# ══════════════════════════════════════════════════════════════════════════════
ws12 = wb_out.create_sheet("#12) AP Trade #2000")
ap = abs(ytd(2000))
ws12.append(["Aurisic A/P Trade #2000"])
ws12["A1"].font = Font(bold=True)
ws12.append(["as of 4-30-25"])
ws12.append([None])
ws12.append([None, "Printed on 05/05/25, Postings Through 4/30/25"])
ws12.append([None, "Vendor", None, None, None, None, "AP Balance"])
flag_cell(
    ws12, 6, 1,
    "AP_TB-1.xlsx referenced in task prompt but not present in reference files. "
    "Vendor-level detail unavailable — showing GL total from trial balance only."
)
ws12.append([None, "Total per GL Trial Balance (acct 0010-2000)", None, None, None, None, ap])
ws12.append([None, "Account Totals:", None, None, None, None, ap])
ws12.cell(7, 7).number_format = FMT_USD
ws12.cell(8, 7).number_format = FMT_USD
print("Tab 12 (AP Trade): done (flagged — AP_TB-1.xlsx missing)")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 13: AR Accruals #1101
# ══════════════════════════════════════════════════════════════════════════════
add_tab("AR_Accrual-1.xlsx", "AR Accrual #1101", "#13) AR Accruals #1101", wb_out)
print("Tab 13 (AR): done")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 14: Payroll Accrual #2200
# Primary: Payroll-1.xlsx; Bonus accrual appended as sub-section
# ══════════════════════════════════════════════════════════════════════════════
ws14 = wb_out.create_sheet("#14) Payroll Accrual #2200")
wb_p = openpyxl.load_workbook(os.path.join(REF, "Payroll-1.xlsx"), data_only=True)
copy_values(wb_p["Payroll Accrual"], ws14)
wb_p.close()
ws14.append([])
ws14.append(["--- Bonus Accrual (AccrBonus-1.xlsx) ---"])
ws14.cell(ws14.max_row, 1).font = Font(bold=True)
wb_b = openpyxl.load_workbook(os.path.join(REF, "AccrBonus-1.xlsx"), data_only=True)
copy_values(wb_b["Bonus Accru #2401"], ws14)
wb_b.close()
print("Tab 14 (Payroll + Bonus): done")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 15: Vendor Rebates #2005
# ══════════════════════════════════════════════════════════════════════════════
add_tab("Rebates-1.xlsx", "Vendor Rebates", "#15) Vendor Rebates #2005", wb_out)
print("Tab 15 (Vendor Rebates): done")

# ══════════════════════════════════════════════════════════════════════════════
# NEW TABS 16–18
# ══════════════════════════════════════════════════════════════════════════════
add_tab("AccrBonus-1.xlsx", "Bonus Accru #2401",            "#16) Bonus Accrual #2401",  wb_out)
add_tab("Accr2011-1.xlsx",  "Aurisic Global Accrual #2011", "#17) Global Accrual #2011", wb_out)
add_tab("AccrMisc-1.xlsx",  "Misc Accruals #2410",          "#18) Misc Accruals #2410",  wb_out)
print("New tabs 16-18 (Accruals): done")

# ══════════════════════════════════════════════════════════════════════════════
# Save
# ══════════════════════════════════════════════════════════════════════════════
wb_out.save(OUT_FILE)
print(f"\nSaved: {OUT_FILE}")
print(f"Tabs ({len(wb_out.sheetnames)}):")
for i, name in enumerate(wb_out.sheetnames):
    print(f"  {i:2}: {name}")

# ══════════════════════════════════════════════════════════════════════════════
# Verification
# ══════════════════════════════════════════════════════════════════════════════
wb_v = openpyxl.load_workbook(OUT_FILE)
assert len(wb_v.sheetnames) >= 18, f"Expected ≥18 tabs, got {len(wb_v.sheetnames)}"
assert "Table of Contents" in wb_v.sheetnames, "TOC missing"
assert "#3a) TB convert 4-30-25" in wb_v.sheetnames, "Tab 3a missing"
for bad in ["3-31-25", "3-25"]:
    for nm in wb_v.sheetnames:
        assert bad not in nm, f"Old date '{bad}' still in tab name: {nm}"
print("\nAll verification checks passed.")
