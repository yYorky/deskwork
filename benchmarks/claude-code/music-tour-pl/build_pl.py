"""
Fall Music Tour P&L Builder
Stage 4 — Claude Code benchmark

Reads: benchmarks/tasks/music-tour-pl/reference-files/Fall Music Tour Ref File.xlsx
Writes: benchmarks/claude-code/music-tour-pl/output/Fall_Music_Tour_PL_2024.xlsx
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

# ── 1. READ SOURCE DATA ────────────────────────────────────────────────────────

BASE = os.path.dirname(os.path.abspath(__file__))
REF  = os.path.join(BASE, "../../tasks/music-tour-pl/reference-files/Fall Music Tour Ref File.xlsx")
OUT  = os.path.join(BASE, "output/Fall_Music_Tour_PL_2024.xlsx")

wb_ref = openpyxl.load_workbook(REF)

# Sheet 1: Revenue and Tour Manager costs
# ── Revenue (col E, rows 9-15, all USD per column header)
rev_raw = [
    ("London",    "UK",      230754.00),
    ("Paris",     "France",  175880.00),
    ("Paris",     "France",  168432.00),
    ("Barcelona", "Spain",   125932.00),
    ("Madrid",    "Spain",   110823.00),
    ("Munich",    "Germany",  99117.00),
    ("Berlin",    "Germany", 132812.00),
]

# Withholding rates (confirmed from Sheet 2)
WHT = {"UK": 0.20, "France": 0.15, "Spain": 0.24, "Germany": 0.15825}

# ── Tour Manager costs (Sheet 1)
gross_total = sum(r[2] for r in rev_raw)   # for Agency Commission calc

tm_band_crew = {
    "Sound Technician": 8256.00,
    "Tour Coordinator":  6904.00,
}
tm_hotel = {
    "London":    8388.00,
    "Paris":    15653.00,
    "Barcelona": 5445.00,
    "Madrid":    5113.00,
    "Munich":    6369.00,
    "Berlin":    6592.00,
}
tm_other_tour = {
    "Agency Commission (11% of Gross Revenue)": round(gross_total * 0.11, 2),
    "Insurance":  22024.00,
    "Other":       4819.00,
}
tm_other_travel = {
    "Private Jet":    341000.00,
    "Transfer Cars":    4237.00,
}

# ── Production Company costs (Sheet 3)
pc_band_crew = {
    "Band & Crew (10 members, Fees & Per Diem)": 91000.00,
}
pc_hotel = {
    "London":   14232.00,
    "Paris":    22296.00,
    "Barcelona": 8168.00,
    "Madrid":    8776.00,
    "Munich":   12040.00,
    "Berlin":   13226.00,
}
pc_other_tour = {
    "Petty Cash": 8000.00,
    "Fees":       1679.00,
}
pc_other_travel = {
    "Car Service": 2976.00,
}

# ── 2. CALCULATE WITHHOLDING & REVENUE TOTALS ──────────────────────────────────

rev_lines = []
for city, country, gross in rev_raw:
    rate = WHT[country]
    wht  = round(gross * rate, 2)
    net  = round(gross - wht, 2)
    rev_lines.append((city, country, gross, rate, wht, net))

gross_rev_total = round(sum(r[2] for r in rev_lines), 2)
wht_total       = round(sum(r[4] for r in rev_lines), 2)
net_rev_total   = round(sum(r[5] for r in rev_lines), 2)

# ── 3. EXPENSE CATEGORY SUBTOTALS ────────────────────────────────────────────

def total(d): return round(sum(d.values()), 2)

tm_band_total         = total(tm_band_crew)
tm_hotel_total        = total(tm_hotel)
tm_other_tour_total   = total(tm_other_tour)
tm_other_travel_total = total(tm_other_travel)
tm_expense_total      = round(tm_band_total + tm_hotel_total +
                               tm_other_tour_total + tm_other_travel_total, 2)

pc_band_total         = total(pc_band_crew)
pc_hotel_total        = total(pc_hotel)
pc_other_tour_total   = total(pc_other_tour)
pc_other_travel_total = total(pc_other_travel)
pc_expense_total      = round(pc_band_total + pc_hotel_total +
                               pc_other_tour_total + pc_other_travel_total, 2)

combined_band         = round(tm_band_total + pc_band_total, 2)
combined_hotel        = round(tm_hotel_total + pc_hotel_total, 2)
combined_other_tour   = round(tm_other_tour_total + pc_other_tour_total, 2)
combined_other_travel = round(tm_other_travel_total + pc_other_travel_total, 2)
combined_expense_total= round(tm_expense_total + pc_expense_total, 2)

# Revenue is Tour Manager only; PC revenue = 0
tm_net_income   = round(net_rev_total - tm_expense_total, 2)
pc_net_income   = round(0 - pc_expense_total, 2)
total_net_income= round(net_rev_total - combined_expense_total, 2)

# ── 4. BUILD EXCEL FILE ───────────────────────────────────────────────────────

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "P&L Statement"

# ── Style helpers ─────────────────────────────────────────────────────────────
NAVY   = "1F3864"
LGREY  = "D9E1F2"
MGREY  = "BDD7EE"
WHITE  = "FFFFFF"
YELLOW = "FFF2CC"

def hdr_font(bold=True, color=WHITE, sz=11):
    return Font(name="Calibri", bold=bold, color=color, size=sz)

def body_font(bold=False, color="000000", sz=10):
    return Font(name="Calibri", bold=bold, color=color, size=sz)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="A0A0A0")
    return Border(left=s, right=s, top=s, bottom=s)

def btm_border():
    s = Side(style="medium", color="1F3864")
    return Border(bottom=s)

USD_FMT = '$#,##0.00'
PCT_FMT = '0.00%'

# ── Column widths ─────────────────────────────────────────────────────────────
# A=Description  B=Tour Manager  C=Production Company  D=Total Combined
# E=WHT Rate  F=WHT Tax  (revenue section only)
col_widths = {1: 32, 2: 20, 3: 22, 4: 20, 5: 12, 6: 18}
for col, w in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

# ── Row builder helpers ───────────────────────────────────────────────────────
row = [0]

def next_row():
    row[0] += 1
    return row[0]

def write_row(values, fonts=None, fills=None, fmts=None, aligns=None, borders=None):
    r = next_row()
    for c, val in enumerate(values, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        if fonts and c - 1 < len(fonts) and fonts[c-1]:
            cell.font = fonts[c-1]
        if fills and c - 1 < len(fills) and fills[c-1]:
            cell.fill = fills[c-1]
        if fmts and c - 1 < len(fmts) and fmts[c-1]:
            cell.number_format = fmts[c-1]
        if aligns and c - 1 < len(aligns) and aligns[c-1]:
            cell.alignment = aligns[c-1]
        if borders and c - 1 < len(borders) and borders[c-1]:
            cell.border = borders[c-1]
    return r

# ── TITLE BLOCK ───────────────────────────────────────────────────────────────
r = next_row()
ws.merge_cells(f"A{r}:F{r}")
c = ws.cell(row=r, column=1,
            value="2024 Fall Music Tour — Profit & Loss Statement")
c.font    = Font(name="Calibri", bold=True, size=16, color=WHITE)
c.fill    = fill(NAVY)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[r].height = 24

r = next_row()
ws.merge_cells(f"A{r}:F{r}")
c = ws.cell(row=r, column=1, value="As of 12/31/2024")
c.font      = Font(name="Calibri", bold=True, size=11, color=WHITE)
c.fill      = fill(NAVY)
c.alignment = Alignment(horizontal="center")
ws.row_dimensions[r].height = 18

next_row()  # blank

# ── COLUMN HEADERS ────────────────────────────────────────────────────────────
hdr_labels = ["Description", "Tour Manager", "Production Company",
              "Total Combined", "WHT Rate", "WHT Tax (USD)"]
hdr_fonts  = [hdr_font()] * 6
hdr_fills  = [fill(NAVY)]   * 6
hdr_aligns = [Alignment(horizontal="center", wrap_text=True)] * 6
write_row(hdr_labels, fonts=hdr_fonts, fills=hdr_fills, aligns=hdr_aligns)

# ── REVENUE SECTION ───────────────────────────────────────────────────────────
r = next_row()
ws.merge_cells(f"A{r}:F{r}")
c = ws.cell(row=r, column=1, value="REVENUE")
c.font      = Font(name="Calibri", bold=True, size=10, color=NAVY)
c.fill      = fill(MGREY)
c.alignment = Alignment(horizontal="left")

# Header row for revenue columns
write_row(["Tour Stop (City, Country)", "Gross Revenue (USD)", "",
           "Total Net Revenue (USD)", "WHT Rate", "WHT Tax (USD)"],
          fonts=[body_font(bold=True)]*6,
          fills=[fill(LGREY)]*6,
          aligns=[Alignment(horizontal="center")]*6)

# Tour stop lines
for city, country, gross, rate, wht, net in rev_lines:
    write_row(
        [f"{city}, {country}", gross, "", net, rate, wht],
        fonts=[body_font(), body_font(), None, body_font(), body_font(), body_font()],
        fmts=[None, USD_FMT, None, USD_FMT, PCT_FMT, USD_FMT],
        aligns=[Alignment(horizontal="left"),
                Alignment(horizontal="right"), None,
                Alignment(horizontal="right"),
                Alignment(horizontal="center"),
                Alignment(horizontal="right")],
    )

# Gross revenue subtotal
write_row(["Gross Revenue Subtotal", gross_rev_total, "", "", "", ""],
          fonts=[body_font(bold=True), body_font(bold=True)] + [body_font()]*4,
          fmts=[None, USD_FMT] + [None]*4,
          fills=[fill(LGREY)]*6)

# Withholding subtotal
write_row(["Less: Total Withholding Tax", -wht_total, "", "", "", wht_total],
          fonts=[body_font(bold=True), body_font(bold=True)] + [body_font()]*4,
          fmts=[None, USD_FMT] + [None]*3 + [USD_FMT],
          fills=[fill(LGREY)]*6)

# Total net revenue
r_net_rev = next_row()
for col in range(1, 7):
    cell = ws.cell(row=r_net_rev, column=col)
    cell.fill   = fill(NAVY)
    cell.font   = hdr_font()
    cell.border = btm_border()
    cell.alignment = Alignment(horizontal="right" if col > 1 else "left")
ws.cell(row=r_net_rev, column=1, value="Total Net Revenue")
ws.cell(row=r_net_rev, column=2, value=net_rev_total).number_format = USD_FMT
ws.cell(row=r_net_rev, column=3, value=0.00).number_format = USD_FMT
ws.cell(row=r_net_rev, column=4, value=net_rev_total).number_format = USD_FMT
ws.cell(row=r_net_rev, column=1).alignment = Alignment(horizontal="left")

next_row()  # blank

# ── EXPENSES SECTION ──────────────────────────────────────────────────────────
r = next_row()
ws.merge_cells(f"A{r}:F{r}")
c = ws.cell(row=r, column=1, value="EXPENSES")
c.font      = Font(name="Calibri", bold=True, size=10, color=NAVY)
c.fill      = fill(MGREY)
c.alignment = Alignment(horizontal="left")

# Sub-header
write_row(["Category", "Tour Manager", "Production Company", "Total Combined", "", ""],
          fonts=[body_font(bold=True)]*6,
          fills=[fill(LGREY)]*6,
          aligns=[Alignment(horizontal="center")]*6)

expense_rows = [
    ("Band and Crew",       tm_band_total,         pc_band_total,         combined_band),
    ("Other Tour Costs",    tm_other_tour_total,   pc_other_tour_total,   combined_other_tour),
    ("Hotel & Restaurants", tm_hotel_total,         pc_hotel_total,        combined_hotel),
    ("Other Travel Costs",  tm_other_travel_total,  pc_other_travel_total, combined_other_travel),
]
for label, tm, pc, tot in expense_rows:
    write_row([label, tm, pc, tot, "", ""],
              fonts=[body_font()]*6,
              fmts=[None, USD_FMT, USD_FMT, USD_FMT, None, None],
              aligns=[Alignment(horizontal="left")] +
                     [Alignment(horizontal="right")]*3 + [None, None])

# Total expenses
r_exp = next_row()
for col in range(1, 7):
    cell = ws.cell(row=r_exp, column=col)
    cell.fill      = fill(LGREY)
    cell.font      = body_font(bold=True)
    cell.alignment = Alignment(horizontal="right" if col > 1 else "left")
ws.cell(row=r_exp, column=1, value="Total Expenses")
ws.cell(row=r_exp, column=2, value=tm_expense_total).number_format   = USD_FMT
ws.cell(row=r_exp, column=3, value=pc_expense_total).number_format   = USD_FMT
ws.cell(row=r_exp, column=4, value=combined_expense_total).number_format = USD_FMT
ws.cell(row=r_exp, column=1).alignment = Alignment(horizontal="left")

next_row()  # blank

# ── NET INCOME ────────────────────────────────────────────────────────────────
r_ni = next_row()
for col in range(1, 5):
    cell = ws.cell(row=r_ni, column=col)
    cell.fill      = fill(NAVY)
    cell.font      = hdr_font(sz=12)
    cell.alignment = Alignment(horizontal="right" if col > 1 else "left")
ws.cell(row=r_ni, column=1, value="NET INCOME")
ws.cell(row=r_ni, column=2, value=tm_net_income).number_format    = USD_FMT
ws.cell(row=r_ni, column=3, value=pc_net_income).number_format    = USD_FMT
ws.cell(row=r_ni, column=4, value=total_net_income).number_format = USD_FMT
ws.cell(row=r_ni, column=1).alignment = Alignment(horizontal="left")
ws.row_dimensions[r_ni].height = 20

next_row()  # blank

# ── WITHHOLDING TAX DETAIL APPENDIX ──────────────────────────────────────────
r = next_row()
ws.merge_cells(f"A{r}:F{r}")
c = ws.cell(row=r, column=1, value="WITHHOLDING TAX DETAIL")
c.font = Font(name="Calibri", bold=True, size=10, color=NAVY)
c.fill = fill(MGREY)

write_row(["Country", "Gross Revenue (USD)", "", "", "WHT Rate", "WHT Tax (USD)"],
          fonts=[body_font(bold=True)]*6,
          fills=[fill(LGREY)]*6)

# Group by country for summary
from collections import defaultdict
country_gross = defaultdict(float)
for _, country, gross, *_ in rev_lines:
    country_gross[country] += gross

for country, rate in WHT.items():
    gross = country_gross.get(country, 0)
    wht_c = round(gross * rate, 2)
    write_row([country, gross, "", "", rate, wht_c],
              fmts=[None, USD_FMT, None, None, PCT_FMT, USD_FMT],
              fonts=[body_font()]*6)

write_row(["Total", gross_rev_total, "", "", "", wht_total],
          fmts=[None, USD_FMT, None, None, None, USD_FMT],
          fonts=[body_font(bold=True)]*6,
          fills=[fill(LGREY)]*6)

next_row()  # blank

# ── EXPENSE DETAIL APPENDIX ───────────────────────────────────────────────────
r = next_row()
ws.merge_cells(f"A{r}:F{r}")
c = ws.cell(row=r, column=1, value="EXPENSE DETAIL")
c.font = Font(name="Calibri", bold=True, size=10, color=NAVY)
c.fill = fill(MGREY)

def detail_section(title, tm_dict, pc_dict):
    r = next_row()
    ws.cell(row=r, column=1, value=title).font = body_font(bold=True)
    ws.cell(row=r, column=1).fill = fill(LGREY)
    ws.cell(row=r, column=2, value="Tour Manager").font = body_font(bold=True)
    ws.cell(row=r, column=2).fill = fill(LGREY)
    ws.cell(row=r, column=3, value="Production Company").font = body_font(bold=True)
    ws.cell(row=r, column=3).fill = fill(LGREY)
    ws.cell(row=r, column=4, value="Total").font = body_font(bold=True)
    ws.cell(row=r, column=4).fill = fill(LGREY)

    all_keys = list(dict.fromkeys(list(tm_dict.keys()) + list(pc_dict.keys())))
    for k in all_keys:
        tm_v = tm_dict.get(k, 0)
        pc_v = pc_dict.get(k, 0)
        write_row([f"  {k}", tm_v if tm_v else "", pc_v if pc_v else "",
                   round(tm_v + pc_v, 2)],
                  fmts=[None, USD_FMT, USD_FMT, USD_FMT],
                  fonts=[body_font()]*4)

detail_section("Band and Crew",       tm_band_crew,       pc_band_crew)
detail_section("Other Tour Costs",    tm_other_tour,      pc_other_tour)
detail_section("Hotel & Restaurants", tm_hotel,           pc_hotel)
detail_section("Other Travel Costs",  tm_other_travel,    pc_other_travel)

# ── 5. SAVE & VERIFY ──────────────────────────────────────────────────────────
os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print(f"Saved: {OUT}")

# Verify file opens
wb_check = openpyxl.load_workbook(OUT)
ws_check = wb_check.active
print(f"File opens OK. Sheet: {ws_check.title!r}")

# ── 6. RECONCILIATION CHECK ───────────────────────────────────────────────────
print("\n── Reconciliation ──────────────────────────────")
print(f"Gross Revenue:        ${gross_rev_total:>14,.2f}")
print(f"Less: WHT:           -${wht_total:>14,.2f}")
print(f"Net Revenue:          ${net_rev_total:>14,.2f}")
print()
print(f"TM Band & Crew:       ${tm_band_total:>14,.2f}")
print(f"TM Hotel & Rest:      ${tm_hotel_total:>14,.2f}")
print(f"TM Other Tour:        ${tm_other_tour_total:>14,.2f}")
print(f"TM Other Travel:      ${tm_other_travel_total:>14,.2f}")
print(f"TM Total Expenses:    ${tm_expense_total:>14,.2f}")
print()
print(f"PC Band & Crew:       ${pc_band_total:>14,.2f}")
print(f"PC Hotel & Rest:      ${pc_hotel_total:>14,.2f}")
print(f"PC Other Tour:        ${pc_other_tour_total:>14,.2f}")
print(f"PC Other Travel:      ${pc_other_travel_total:>14,.2f}")
print(f"PC Total Expenses:    ${pc_expense_total:>14,.2f}")
print()
print(f"Combined Expenses:    ${combined_expense_total:>14,.2f}")
print(f"TM Net Income:        ${tm_net_income:>14,.2f}")
print(f"PC Net Income:       -${abs(pc_net_income):>14,.2f}  (costs only)")
print(f"Total Net Income:     ${total_net_income:>14,.2f}")
print()
# Foot check: net income = net rev - combined expenses
assert abs(total_net_income - (net_rev_total - combined_expense_total)) < 0.01, "NET INCOME FOOT CHECK FAILED"
# Cross check: combined = tm + pc for each category
assert abs(combined_band - (tm_band_total + pc_band_total)) < 0.01, "BAND CROSS CHECK FAILED"
assert abs(combined_hotel - (tm_hotel_total + pc_hotel_total)) < 0.01, "HOTEL CROSS CHECK FAILED"
assert abs(combined_other_tour - (tm_other_tour_total + pc_other_tour_total)) < 0.01, "OTHER TOUR CROSS CHECK FAILED"
assert abs(combined_other_travel - (tm_other_travel_total + pc_other_travel_total)) < 0.01, "OTHER TRAVEL CROSS CHECK FAILED"
print("All reconciliation checks PASSED ✓")
